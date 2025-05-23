import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from fuzzywuzzy import fuzz
import io
import re

def load_excel_file(file):
    """
    Загружает Excel файл и возвращает объект рабочей книги и список листов
    
    Args:
        file: Загруженный файл в формате BytesIO
        
    Returns:
        Tuple: (workbook, list_of_sheets)
    """
    workbook = openpyxl.load_workbook(file, data_only=False)
    sheets = workbook.sheetnames
    
    return workbook, sheets

def save_excel_file(workbook):
    """
    Сохраняет Excel файл и возвращает BytesIO объект
    
    Args:
        workbook: Объект рабочей книги openpyxl
        
    Returns:
        BytesIO: Объект байтового потока для скачивания
    """
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output

def map_columns_automatically(source_columns, target_columns, threshold=70):
    """
    Автоматически сопоставляет колонки на основе схожести названий
    
    Args:
        source_columns: Список колонок исходной таблицы
        target_columns: Список колонок целевой таблицы
        threshold: Порог схожести для сопоставления (0-100)
        
    Returns:
        Dict: Словарь соответствия {source_column: target_column}
    """
    mapping = {}
    used_target_columns = set()
    
    # Улучшенная нормализация имен колонок
    def normalize_column_name(col_name):
        if not isinstance(col_name, str):
            return str(col_name).lower()
        
        # Сохраняем исходное имя колонки для дальнейшего анализа
        original_name = col_name
        
        # Удаляем специальные маркеры из заголовков маркетплейсов: звездочки, восклицательные знаки и т.д.
        normalized = re.sub(r'[*!№\+]', '', col_name)
        
        # Удаляем спецсимволы и лишние пробелы
        normalized = re.sub(r'[^\w\s\-\.]', ' ', normalized)
        normalized = re.sub(r'\s+', ' ', normalized).strip().lower()
        
        # Обрабатываем многострочные заголовки (перенос строки заменяем на пробел)
        normalized = re.sub(r'\n', ' ', normalized)
        
        # Удаляем распространенные суффиксы
        common_suffixes = [
            " товара", " продукта", " позиции", " изделия", " шт", " г", " кг", " мл", " л",
            " см", " мм", " м", " руб", " rub", " ₽", " %", " руб."
        ]
        for suffix in common_suffixes:
            if normalized.endswith(suffix.lower()):
                normalized = normalized[:-len(suffix)]
        
        # Удаляем общие префиксы и пояснения в скобках
        normalized = re.sub(r'\([^)]*\)', '', normalized).strip()
        common_prefixes = [
            "код ", "номер ", "ид ", "тип ", "название ", "наименование ", "цена ", "стоимость ",
            "размер ", "вес ", "масса ", "ширина ", "высота ", "глубина ", "длина "
        ]
        for prefix in common_prefixes:
            if normalized.startswith(prefix):
                normalized = normalized[len(prefix):]
        
        # Обрабатываем заголовки с единицами измерения через запятую
        normalized = re.sub(r',\s*(шт|г|кг|мл|л|см|мм|м|руб|rub|₽|%)$', '', normalized)
        
        # Преобразуем сокращения в полные формы
        abbreviations = {
            "артик": "артикул",
            "наим": "название",
            "наимен": "название",
            "описан": "описание",
            "кол-во": "количество",
            "кол во": "количество",
            "колво": "количество",
            "кол": "количество",
            "хар-ки": "характеристики",
            "хар ки": "характеристики",
            "харки": "характеристики",
            "хар-ка": "характеристика",
            "хар ка": "характеристика",
            "харка": "характеристика",
            "спец": "спецификация",
            "габар": "габариты",
            "разм": "размер",
            "фото": "изображение",
            "изобр": "изображение",
            "изобрa": "изображение",
            "картин": "изображение",
            "шир": "ширина",
            "дл": "длина",
            "выс": "высота",
            "глуб": "глубина"
        }
        for abbr, full in abbreviations.items():
            if normalized == abbr or normalized.startswith(abbr + " "):
                normalized = normalized.replace(abbr, full, 1)
                break
                
        # Стандартизация распространенных названий колонок для маркетплейсов
        marketplace_columns = {
            # Ключевые атрибуты
            "артикул": ["арт", "артик", "код товара", "номер артикула", "skuмагазина", "sku", "id товара", 
                      "код позиции", "wb sku", "артикул wb", "ozon id", "артикул продавца"],
            "название": ["наименование", "имя", "наимен", "назв", "имя товара", "заголовок", "title", 
                       "наименование товара", "название товара", "наименование позиции", "полное название"],
            "цена": ["стоимость", "розн цена", "цена продажи", "price", "розничная цена", "прайс", 
                    "цена товара", "цена со скидкой", "розничная", "цена розничная", "руб"],
            "описание": ["описание товара", "полное описание", "detail", "детальное описание", "description", 
                        "расширенное описание", "контент", "content", "информация о товаре", "товар описание"],
            "категория": ["раздел", "группа", "группа товаров", "category", "тип товара", "тип изделия", 
                        "категория товара", "родительская категория", "товарная категория", "предметная группа"],
            "бренд": ["брэнд", "марка", "производитель", "brand", "изготовитель", "торговая марка", "тм", 
                     "товарный знак", "компания производитель", "марка производитель"],
            
            # Габаритные характеристики
            "вес": ["масса", "вес товара", "вес в упаковке", "вес без упаковки", "вес брутто", "вес нетто", 
                   "weight", "масса товара", "масса в упаковке", "объемный вес"],
            "ширина": ["width", "ширина товара", "ширина упаковки", "ширина изделия", "ширина габарит",
                      "ширина в упаковке", "ширина без упаковки", "габариты ширина"],
            "высота": ["height", "высота товара", "высота упаковки", "высота изделия", "высота габарит",
                      "высота в упаковке", "высота без упаковки", "габариты высота"],
            "длина": ["length", "глубина", "длина товара", "длина упаковки", "длина изделия", "длина габарит",
                     "длина в упаковке", "длина без упаковки", "габариты длина", "глубина габарит"],
            
            # Логистика и наличие
            "количество": ["кол-во", "остаток", "остатки", "наличие", "колво", "qty", "quantity", 
                          "количество штук", "доступное количество", "количество в наличии"],
            "баркод": ["штрихкод", "шк", "баркод товара", "ean", "ean13", "gtin", "upc", "код товара", 
                      "штрих код", "barcode", "штрихкод товара"],
            
            # Дополнительные атрибуты
            "материал": ["состав", "материал изготовления", "материал товара", "материал изделия", 
                        "основной материал", "material", "ткань", "основа", "сырье"],
            "цвет": ["color", "расцветка", "цвет товара", "цвет изделия", "основной цвет", 
                    "цветовой тон", "оттенок", "цвет и оттенок", "цветовое решение"],
            "размер": ["габариты", "size", "размерный ряд", "размер товара", "размер изделия", 
                      "линейные размеры", "типоразмер", "размерность", "габаритные размеры"]
        }
        
        # Проверяем, соответствует ли нормализованное имя одному из стандартных имён
        for standard, aliases in marketplace_columns.items():
            if normalized == standard or normalized in aliases:
                return standard
                
        # Для оригинальных имен колонок с маркерами обязательности (звездочка и др.)
        # добавляем повышенный приоритет для ключевых атрибутов
        if '*' in original_name or '!' in original_name:
            for key_field, aliases in {'артикул': marketplace_columns['артикул'], 
                                      'название': marketplace_columns['название'], 
                                      'цена': marketplace_columns['цена']}.items():
                for alias in aliases:
                    if alias in normalized or normalized in alias:
                        return key_field
                        
        # Если колонка содержит цифры (например, Артикул1, Артикул2), очищаем от них
        normalized = re.sub(r'\d+$', '', normalized).strip()
                
        return normalized
    
    # Сначала нормализуем все колонки
    source_normalized = {col: normalize_column_name(col) for col in source_columns}
    target_normalized = {col: normalize_column_name(col) for col in target_columns}
    
    # Создаем индекс нормализованных значений для более быстрого поиска
    target_by_norm = {}
    for tgt_col, tgt_norm in target_normalized.items():
        if tgt_norm not in target_by_norm:
            target_by_norm[tgt_norm] = []
        target_by_norm[tgt_norm].append(tgt_col)
    
    # Сначала сопоставляем ключевые колонки (артикул, название, цена)
    key_columns = {'артикул': [], 'название': [], 'цена': []}
    
    # Находим ключевые колонки в исходной таблице
    for src_col, src_norm in source_normalized.items():
        if src_norm in key_columns:
            key_columns[src_norm].append(src_col)
    
    # Сопоставляем ключевые колонки первыми
    for key, src_cols in key_columns.items():
        if src_cols and key in target_by_norm:
            for src_col in src_cols:
                if src_col not in mapping:
                    for tgt_col in target_by_norm[key]:
                        if tgt_col not in used_target_columns:
                            mapping[src_col] = tgt_col
                            used_target_columns.add(tgt_col)
                            break
    
    # Сначала ищем точные совпадения нормализованных имён
    for src_col, src_norm in source_normalized.items():
        if src_col not in mapping and src_norm in target_by_norm:
            for tgt_col in target_by_norm[src_norm]:
                if tgt_col not in used_target_columns:
                    mapping[src_col] = tgt_col
                    used_target_columns.add(tgt_col)
                    break
    
    # Затем ищем по полному совпадению оригинальных названий
    for src_col in source_columns:
        if src_col not in mapping:
            for tgt_col in target_columns:
                if src_col == tgt_col and tgt_col not in used_target_columns:
                    mapping[src_col] = tgt_col
                    used_target_columns.add(tgt_col)
                    break
    
    # Ищем по частичному совпадению после нормализации
    for src_col, src_norm in source_normalized.items():
        if src_col not in mapping:
            best_match = None
            longest_common = 0
            
            for tgt_col, tgt_norm in target_normalized.items():
                if tgt_col not in used_target_columns:
                    # Если один является вложенным в другой
                    if (src_norm in tgt_norm or tgt_norm in src_norm) and len(min(src_norm, tgt_norm, key=len)) > 3:
                        common_length = len(min(src_norm, tgt_norm, key=len))
                        if common_length > longest_common:
                            longest_common = common_length
                            best_match = tgt_col
            
            if best_match:
                mapping[src_col] = best_match
                used_target_columns.add(best_match)
    
    # Используем нечеткое сопоставление с улучшенным подходом
    for src_col in source_columns:
        if src_col not in mapping:
            best_match = None
            best_score = 0
            
            # Получаем нормализованное имя исходной колонки
            src_norm = source_normalized[src_col]
            
            for tgt_col in target_columns:
                if tgt_col not in used_target_columns:
                    # Получаем нормализованное имя целевой колонки
                    tgt_norm = target_normalized[tgt_col]
                    
                    # Рассчитываем несколько метрик для сравнения
                    metrics = [
                        fuzz.ratio(src_norm, tgt_norm),                      # Общее сходство строк
                        fuzz.partial_ratio(src_norm, tgt_norm),              # Частичное сходство
                        fuzz.token_sort_ratio(src_norm, tgt_norm),           # Схожесть с сортировкой токенов
                        fuzz.token_set_ratio(src_norm, tgt_norm),            # Схожесть набора токенов
                        fuzz.partial_token_sort_ratio(src_norm, tgt_norm)    # Частичная схожесть с сортировкой
                    ]
                    
                    # Для коротких строк повышаем значимость точного совпадения
                    if len(src_norm) < 6 or len(tgt_norm) < 6:
                        exact_match_bonus = 30 if src_norm[0:3] == tgt_norm[0:3] else 0
                    else:
                        exact_match_bonus = 0
                    
                    # Увеличиваем значимость совпадения начала строки
                    prefix_bonus = 0
                    min_len = min(len(src_norm), len(tgt_norm))
                    if min_len >= 3:
                        prefix_len = 0
                        for i in range(min_len):
                            if src_norm[i] == tgt_norm[i]:
                                prefix_len += 1
                            else:
                                break
                        
                        if prefix_len >= 3:
                            prefix_bonus = prefix_len * 5  # 5 бонусных баллов за каждый совпадающий символ в начале
                    
                    # Рассчитываем взвешенное значение схожести
                    weighted_score = max(metrics) + exact_match_bonus + prefix_bonus
                    
                    if weighted_score > best_score and weighted_score >= threshold:
                        best_score = weighted_score
                        best_match = tgt_col
            
            if best_match:
                mapping[src_col] = best_match
                used_target_columns.add(best_match)
    
    return mapping

def transfer_data_between_tables(source_df, target_workbook, target_sheet_name, column_mapping, target_header_row=1):
    """
    Переносит данные из исходного DataFrame в целевую таблицу, сохраняя форматирование
    
    Args:
        source_df: DataFrame с исходными данными
        target_workbook: Объект целевой рабочей книги openpyxl
        target_sheet_name: Имя целевого листа
        column_mapping: Словарь соответствия колонок {source_column: target_column}
        target_header_row: Номер строки с заголовками в целевой таблице (по умолчанию 1)
        
    Returns:
        Объект рабочей книги openpyxl с обновленными данными
    """
    # Получаем целевой лист
    target_sheet = target_workbook[target_sheet_name]
    
    # Находим индексы колонок в целевой таблице
    header_row = target_header_row  # Используем переданный номер строки с заголовками
    header_cells = list(target_sheet.rows)[header_row - 1]
    
    target_column_indices = {}
    for cell in header_cells:
        if cell.value:
            target_column_indices[str(cell.value)] = cell.column
    
    # Сохраняем информацию о форматировании и подзаголовках в целевой таблице
    style_info = {}
    subheader_info = {}
    first_data_row = header_row + 1  # Первая строка с данными (после заголовка)
    
    # Проверяем наличие подзаголовков или дополнительной информации непосредственно под заголовками
    has_subheaders = False
    subheader_row = target_sheet.cell(row=first_data_row, column=1).value
    if subheader_row is not None and isinstance(subheader_row, str) and not any(char.isdigit() for char in subheader_row):
        has_subheaders = True
        # Сохраняем подзаголовки для каждой колонки
        for col_name, col_idx in target_column_indices.items():
            subheader_cell = target_sheet.cell(row=first_data_row, column=col_idx)
            if subheader_cell.value:
                subheader_info[col_name] = {
                    'value': subheader_cell.value,
                    'font': subheader_cell.font.copy() if subheader_cell.font else None,
                    'fill': subheader_cell.fill.copy() if subheader_cell.fill else None,
                    'border': subheader_cell.border.copy() if subheader_cell.border else None,
                    'alignment': subheader_cell.alignment.copy() if subheader_cell.alignment else None,
                    'number_format': subheader_cell.number_format,
                    'protection': subheader_cell.protection.copy() if subheader_cell.protection else None,
                }
        # Смещаем первую строку с данными, если есть подзаголовки
        first_data_row += 1
    
    # Сохраняем все существующие данные в целевой таблице для анализа
    existing_data = {}
    max_row = min(target_sheet.max_row, header_row + 20)  # Ограничиваем для производительности
    
    for row_idx in range(header_row + 1, max_row + 1):
        row_data = {}
        for col_name, col_idx in target_column_indices.items():
            cell = target_sheet.cell(row=row_idx, column=col_idx)
            row_data[col_name] = {
                'value': cell.value,
                'font': cell.font.copy() if cell.font else None,
                'fill': cell.fill.copy() if cell.fill else None,
                'border': cell.border.copy() if cell.border else None,
                'alignment': cell.alignment.copy() if cell.alignment else None,
                'number_format': cell.number_format,
                'protection': cell.protection.copy() if cell.protection else None,
            }
        existing_data[row_idx] = row_data
    
    # Выявляем строки с пояснениями/подсказками
    hint_rows = []
    for row_idx, row_data in existing_data.items():
        # Проверка 1: Строка непосредственно после заголовка может быть подсказкой
        if row_idx == header_row + 1:
            # Проверяем, достаточно ли похожа строка на подсказку
            text_only = True
            hint_pattern = False
            long_text_count = 0
            
            for col_name, cell_info in row_data.items():
                val = cell_info['value']
                if val is not None:
                    val_str = str(val) if not isinstance(val, str) else val
                    if len(val_str) > 15:  # Длинный текст может быть подсказкой
                        long_text_count += 1
                    if isinstance(val, (int, float)) or (isinstance(val, str) and val.isdigit()):
                        text_only = False
                    # Проверяем если текст содержит типичные слова-маркеры подсказок
                    hint_markers = ["заполнить", "заполняйте", "указать", "указывать", "используйте", 
                                   "только для", "вводите", "укажите", "обязательно", "не более"]
                    if isinstance(val, str) and any(marker in val.lower() for marker in hint_markers):
                        hint_pattern = True
            
            # Если это первая строка и у неё есть признаки подсказок, признаем её подсказкой
            if (text_only and long_text_count > 0) or hint_pattern:
                hint_rows.append(row_idx)
                continue
        
        # Проверка 2: Стандартное обнаружение строк с подсказками (текстовые строки без чисел)
        text_only = True
        has_content = False
        long_text_found = False
        
        for col_name, cell_info in row_data.items():
            val = cell_info['value']
            if val is not None:
                has_content = True
                if isinstance(val, str) and len(val) > 15:
                    long_text_found = True
                
                # Если значение содержит цифры и не выглядит как подсказка, то это не подсказка
                if isinstance(val, (int, float)) or (isinstance(val, str) and any(c.isdigit() for c in val)):
                    # Проверяем, может ли строка с цифрами всё же быть подсказкой
                    # (например, "минимум 5 символов" или "не более 100 знаков")
                    if not isinstance(val, str) or not any(marker in val.lower() for marker in 
                                                          ["минимум", "максимум", "не более", "не менее", 
                                                           "до", "от", "символов", "знаков"]):
                        text_only = False
                        break
        
        # Строка считается подсказкой если: 
        # - содержит только текст
        # - имеет содержимое (не пустая)
        # - содержит достаточно длинный текст (более вероятно для подсказок)
        if has_content and text_only and long_text_found:
            hint_rows.append(row_idx)
    
    # Определяем последнюю строку с подсказками
    last_hint_row = max(hint_rows) if hint_rows else header_row
    
    # Сначала определим строку, которая точно содержит данные, а не подсказки
    # Ищем первую строку с числовыми данными после заголовков
    data_sample_row = None
    for row_idx in range(header_row + 1, min(target_sheet.max_row + 1, header_row + 20)):
        has_numeric_data = False
        for col_name, col_idx in target_column_indices.items():
            cell_value = target_sheet.cell(row=row_idx, column=col_idx).value
            if isinstance(cell_value, (int, float)) or (isinstance(cell_value, str) and any(c.isdigit() for c in cell_value)):
                has_numeric_data = True
                break
        
        if has_numeric_data and row_idx not in hint_rows:
            data_sample_row = row_idx
            break
    
    # Если не нашли строку с данными, берем последнюю строку после подсказок
    if data_sample_row is None:
        data_sample_row = last_hint_row + 1
        
    # Сохраняем стили форматирования из найденной строки с данными (не из подсказок)
    for col_name, col_idx in target_column_indices.items():
        # Берем ячейку из строки с данными для сохранения стиля
        template_cell = target_sheet.cell(row=data_sample_row, column=col_idx)
        style_info[col_name] = {
            'font': template_cell.font.copy() if template_cell.font else None,
            'fill': template_cell.fill.copy() if template_cell.fill else None,
            'border': template_cell.border.copy() if template_cell.border else None,
            'alignment': template_cell.alignment.copy() if template_cell.alignment else None,
            'number_format': template_cell.number_format,
            'protection': template_cell.protection.copy() if template_cell.protection else None,
        }
    
    # Восстанавливаем все подсказки в ячейках из полученной ранее информации
    for row_idx in hint_rows:
        for col_name, col_idx in target_column_indices.items():
            if row_idx in existing_data and col_name in existing_data[row_idx]:
                cell_info = existing_data[row_idx][col_name]
                cell = target_sheet.cell(row=row_idx, column=col_idx)
                cell.value = cell_info['value']
                
                # Восстанавливаем форматирование
                if cell_info['font']: cell.font = cell_info['font']
                if cell_info['fill']: cell.fill = cell_info['fill']
                if cell_info['border']: cell.border = cell_info['border']
                if cell_info['alignment']: cell.alignment = cell_info['alignment']
                if cell_info['number_format']: cell.number_format = cell_info['number_format']
                if cell_info['protection']: cell.protection = cell_info['protection']
    
    # Удаляем только строки с числовыми данными, сохраняя подсказки
    for row_idx in range(header_row + 1, target_sheet.max_row + 1):
        if row_idx not in hint_rows:
            for col_name, col_idx in target_column_indices.items():
                target_sheet.cell(row=row_idx, column=col_idx).value = None
    
    # Пропускаем первые строки в исходной таблице, если они содержат подзаголовки
    data_start_idx = 0
    first_row_is_numeric = False
    
    if len(source_df) > 0:
        # Проверяем, является ли первая строка числовой (для Pandas индексация с 0)
        first_row = source_df.iloc[0]
        numeric_values = 0
        string_descriptors = 0
        
        for col in source_df.columns:
            val = first_row[col]
            if isinstance(val, (int, float)) and not pd.isna(val):
                numeric_values += 1
            elif isinstance(val, str) and not any(c.isdigit() for c in val):
                string_descriptors += 1
        
        # Если в первой строке больше нечисловых описательных значений, это может быть подзаголовок
        first_row_is_numeric = numeric_values > string_descriptors
        
        if not first_row_is_numeric:
            data_start_idx = 1  # Пропускаем первую строку при переносе данных
    
    # Определяем, где начинать вставку данных в целевой таблице
    # Должно быть после всех строк с подсказками
    data_insertion_start = last_hint_row + 1
    
    # Для каждой строки в source_df (начиная с индекса data_start_idx) 
    # создаем набор значений для записи в целевую таблицу
    for idx, source_row in source_df.iloc[data_start_idx:].iterrows():
        # Вычисляем индекс строки в целевой таблице
        row_offset = idx - data_start_idx if data_start_idx > 0 else idx
        target_row_idx = row_offset + data_insertion_start
        
        # Переносим данные в соответствии с маппингом
        for source_col, target_col in column_mapping.items():
            if target_col in target_column_indices:
                target_col_idx = target_column_indices[target_col]
                
                # Получаем значение из исходной таблицы
                value = source_row.get(source_col)
                
                # Обработка типов данных для обеспечения совместимости
                if value is not None:
                    # Преобразуем все значения в строки для большей совместимости
                    if isinstance(value, (int, float)):
                        # Числовые значения не конвертируем, оставляем как есть
                        pass
                    elif not isinstance(value, str):
                        # Для других типов данных делаем строковое представление
                        value = str(value)
                
                # Записываем значение в целевую таблицу
                cell = target_sheet.cell(row=target_row_idx, column=target_col_idx)
                cell.value = value
                
                # Применяем сохраненное форматирование из образца данных (не из подсказок)
                if target_col in style_info:
                    cell_style = style_info[target_col]
                    if cell_style['font']: cell.font = cell_style['font']
                    if cell_style['fill']: cell.fill = cell_style['fill']
                    if cell_style['border']: cell.border = cell_style['border']
                    if cell_style['alignment']: cell.alignment = cell_style['alignment']
                    if cell_style['number_format']: cell.number_format = cell_style['number_format']
                    if cell_style['protection']: cell.protection = cell_style['protection']
    
    # Восстанавливаем подзаголовки в целевой таблице, если они были
    if has_subheaders:
        for col_name, col_idx in target_column_indices.items():
            if col_name in subheader_info:
                subheader_cell = target_sheet.cell(row=header_row + 1, column=col_idx)
                subheader_cell.value = subheader_info[col_name]['value']
                
                # Восстанавливаем форматирование подзаголовка
                subheader_style = subheader_info[col_name]
                if subheader_style['font']: subheader_cell.font = subheader_style['font']
                if subheader_style['fill']: subheader_cell.fill = subheader_style['fill']
                if subheader_style['border']: subheader_cell.border = subheader_style['border']
                if subheader_style['alignment']: subheader_cell.alignment = subheader_style['alignment']
                if subheader_style['number_format']: subheader_cell.number_format = subheader_style['number_format']
                if subheader_style['protection']: subheader_cell.protection = subheader_style['protection']
    
    return target_workbook

def preview_data(source_df, target_df, column_mapping):
    """
    Создает предварительный просмотр того, как данные будут выглядеть после переноса
    
    Args:
        source_df: DataFrame с исходными данными
        target_df: DataFrame целевой таблицы
        column_mapping: Словарь соответствия колонок {source_column: target_column}
        
    Returns:
        DataFrame: DataFrame с предварительным просмотром
    """
    # Создаем копию целевого DataFrame
    preview_df = target_df.copy()
    
    # Проверяем наличие подзаголовков в исходной таблице
    has_source_subheaders = False
    if len(source_df) > 0:
        first_row = source_df.iloc[0]
        numeric_values = 0
        string_descriptors = 0
        
        for col in source_df.columns:
            val = first_row[col]
            if isinstance(val, (int, float)) and not pd.isna(val):
                numeric_values += 1
            elif isinstance(val, str) and not any(c.isdigit() for c in val):
                string_descriptors += 1
        
        # Если в первой строке больше нечисловых описательных значений, это может быть подзаголовок
        has_source_subheaders = string_descriptors > numeric_values
    
    # Фильтруем данные, пропуская подзаголовки если они есть
    data_start_idx = 1 if has_source_subheaders else 0
    filtered_source_df = source_df.iloc[data_start_idx:].copy() if len(source_df) > data_start_idx else source_df.copy()
    
    # Создаем соответствие между колонками источника и целевой таблицы
    col_pairs = [(src, tgt) for src, tgt in column_mapping.items() if tgt in preview_df.columns]
    
    # Проверяем наличие подзаголовков в целевой таблице
    has_target_subheaders = False
    if len(preview_df) > 0:
        first_row = preview_df.iloc[0]
        numeric_values = 0
        string_descriptors = 0
        
        for col in preview_df.columns:
            val = first_row[col]
            if isinstance(val, (int, float)) and not pd.isna(val):
                numeric_values += 1
            elif isinstance(val, str) and not any(c.isdigit() for c in val) and not pd.isna(val):
                string_descriptors += 1
        
        has_target_subheaders = string_descriptors > numeric_values
    
    # Создаем новый DataFrame для предпросмотра
    result_columns = preview_df.columns
    result_df = pd.DataFrame(columns=result_columns)
    
    # Сохраняем подзаголовки из целевой таблицы, если они есть
    if has_target_subheaders and len(preview_df) > 0:
        result_df = pd.concat([result_df, preview_df.iloc[[0]]], ignore_index=True)
        target_data_start = 1
    else:
        target_data_start = 0
    
    # Копируем данные из исходной таблицы в соответствующие колонки целевой
    for idx, source_row in filtered_source_df.iterrows():
        new_row = {col: None for col in result_columns}
        
        for src, tgt in col_pairs:
            if src in filtered_source_df.columns:
                value = source_row.get(src)
                
                # Обработка значения для совместимости
                if value is not None:
                    # Если это число, оставляем как есть
                    if isinstance(value, (int, float)):
                        pass
                    # Для других типов - преобразуем в строку
                    elif pd.isna(value):
                        value = None
                    elif not isinstance(value, str):
                        value = str(value)
                
                new_row[tgt] = value
        
        # Добавляем строку с данными к результату
        result_df = pd.concat([result_df, pd.DataFrame([new_row])], ignore_index=True)
    
    # Преобразуем все колонки в строковый тип для предотвращения проблем при отображении
    for col in result_df.columns:
        # Сначала конвертируем в строки те значения, что могут вызывать проблемы
        # Используем метод, который будет работать с разными типами данных
        result_df[col] = result_df[col].apply(lambda x: str(x) if x is not None and not pd.isna(x) else None)
    
    return result_df
